import json
import fitz  # PyMuPDF
from openpyxl import load_workbook
from pathlib import Path

FIELDS_JSON = Path("data/mappings/fields_list.json")
EXCEL_SHEET = Path("data/excel/template.xlsx")
BLANK_PDF = Path("data/templates/ticket_blank.pdf")
OUTPUT_DIR = Path("output/tickets")


def generate():
    with open(FIELDS_JSON, "r") as f:
        fields = json.load(f)

    wb = load_workbook(EXCEL_SHEET)
    sheet = wb.active
    headers = [c.value for c in sheet[1]]

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        doc = fitz.open(BLANK_PDF)
        page = doc[0]

        for field in fields:
            key = field["suggested_field"]
            if key not in data:
                continue

            text = str(data[key])
            x, y = field["x"], field["y"]

            page.insert_text((x, y), text, fontsize=10, fontname="helv")

        ticket_id = data.get("ticket_id", "ticket")
        save_path = OUTPUT_DIR / f"{ticket_id}.pdf"

        doc.save(save_path)
        doc.close()
        print("[✔] Generated:", save_path)


if __name__ == "__main__":
    generate()
